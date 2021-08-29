#!/usr/bin/env python
#-*- coding:utf-8 -*-

# COSMOS Project.2015 https://cosmosproject2015.tistory.com/ 
# cosmosproject15@gmail.com

## LICENSE : GPL 2
# Maintainer : cosmosproject15@gmail.com
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.


import multiprocessing
import time
import random
import re
import sys
import subprocess
import linecache
import os
import operator
import pandas as pd
from random import shuffle
from openpyxl import load_workbook
import collections
import warnings


nnum=int(sys.argv[1])
bnum=int(sys.argv[2])
amount=int(sys.argv[3])
avera=int(sys.argv[4])
half=int(sys.argv[5])
date=str(sys.argv[6])
camel=str(sys.argv[7])
test=str(sys.argv[8])

cc=os.cpu_count()
stime=time.time()

snum=nnum-bnum
value=2
warnings.simplefilter("ignore")

load_wb = load_workbook("/tmp/randlot_data.xlsx", data_only=True)
sheet=load_wb.active
warnings.simplefilter("default")
warnings.simplefilter("ignore")

sheet.title = "1"
load_ws = load_wb["1"]


# 오리지날 리스트
olist=list(range(1,46))
# 적용 소수 설계

n=45
a = [False,False] + [True]*(n-1)
primes=[]

for i in range(2,n+1):
	if a[i]:
		primes.append(i)
		for j in range(2*i, n+1, i):
			a[j] = False

# 기초 난수 설계


ranlists=[]

for i in range(2):
	for i in range(45):
		ia=i+1
		ranlists.append(ia)
for i in range(cc):
 

	shuffle(ranlists)
	


# 추가 적용 함수 
def add(num):
	ranlists.append(num)

	
dnotlist=olist
dnot_dict={}

# 미출현 dict 작성 
for i in range(45):
	ai=i+1
	dnot_dict[ai]=0
	

for i in range(snum):
	a=4+i
	datasel1="B"+str(a)
	data=load_ws[datasel1].value
	data1=int(data)

	datasel2="N"+str(a)
	data=load_ws[datasel2].value
	data2=int(data)

	datasel3="O"+str(a)
	data=load_ws[datasel3].value
	data3=int(data)


	datasel4="P"+str(a)
	data=load_ws[datasel4].value
	data4=int(data)

	datasel5="Q"+str(a)
	data=load_ws[datasel5].value
	data5=int(data)


	datasel6="R"+str(a)
	data=load_ws[datasel6].value
	data6=int(data)

	datasel7="S"+str(a)
	data=load_ws[datasel7].value
	data7=int(data)
	
	nowlist=[data2]+[data3]+[data4]+[data5]+[data6]+[data7]
	
	for i in range(6):
		ia=nowlist[i]
		ranlists.append(ia)

	for i in range(5):
		ia=i+1
		nn=nowlist[ia]
		for i in range(2):
			
			add(nn)
	
	if a ==4:
		first_list=nowlist
		print("\n이전 회차 : ", data1,"의 당첨번호는 ",first_list,"입니다.")
		
# 범위 수 적용
		onums=[]
		def drange(arg):
			seldata=arg-5
			for i in range(10):
				onum=seldata+i
				if 46 > onum > 0:
					add(onum)
					onums.append(onum)
		for i in range(3):
			
			drange(data2)
			drange(data3)
			drange(data4)
			drange(data5)
			drange(data6)
			drange(data7)

			# 홀짝 분석

			h=0
			z=0
			if data2%2 ==1:
				h=h+1
			else:
				z=z+1

			if data3%2 ==1:
				h=h+1
			else:
				z=z+1
			if data4%2 ==1:
				h=h+1
			else:
				z=z+1
			if data5%2 ==1:
				h=h+1
			else:
				z=z+1
			if data6%2 ==1:
				h=h+1
			else:
				z=z+1
			if data7%2 ==1:
				h=h+1
			else:
				z=z+1
			if h > 3  :
				for i in range(45):
					sel=1+i
					if sel%2 == 1:
						pass
					else:
						add(sel)
			if z > 3:
				for i in range(45):
					sel=1+i
					if sel%2 == 1:
						add(sel)
			
			# 소수 분석 
			sosuc=set(nowlist) & set(primes)
			
			if len(sosuc) > 3:
				for i in range(45):
					sel=1+i
					c=primes.count(sel)
					if c == 0:
						add(sel)
			if len(sosuc) < 0:
				for i in range(45):
					sel=1+i
					c=primes.count(sel)
					if c == 1:
						add(sel)
			# 번호합 분석
			hap=data2+data3+data4+data5+data6+data7
			if hap > 128 :
				for i in range(22):
					sel=1+i
					for i in range(3):
						add(sel)


			if hap < 128 :
				for i in range(23):
					sel=23+i
					for i in range(3):
						add(sel)

	# 연속 미출현 번호 연산

	dnot=set(dnotlist) - set(nowlist)

	dnotlist=list(dnot)
	
	for i in range(2):
	 
		for i in range(len(dnotlist)):
			dnot_num=dnotlist[i]
		
			dnot_dict[dnot_num]=dnot_dict[dnot_num]+1
		
			add(dnot_num)


onums_dict={}
for i in range(45):
	ia=i+1
	if ia in onums:
		count_ia=onums.count(ia)
		onums_dict[ia]=count_ia

sorted_onums=sorted(onums_dict.items(), key=operator.itemgetter(1), reverse=True)

print("\n범위수 목록" ,sorted_onums)


# 프린트
print("\n기본 분석 현황")
if h > 3  :
	print("이전회차는 홀수가 더 많았습니다.")

if z > 3:
	print("\n이전회차는 짝수가 더 많았습니다.")
	
if z == 3:
	print("\n이전 회차는 홀짝이 공동 출현 했습니다.")

print("\n소수 출현 횟수 :", len(sosuc), sosuc)

if len(sosuc) == 3:
	print("\n소수와 비소수가 공동 출현 했습니다.")


if len(sosuc) > 3:
	print("\n이전 회차는 소수가 더 많이 출현 했었습니다.")
	
if len(sosuc) < 3:
	print("\n이전 회차는 소수가 더 적게 출현 했습니다.")
	
	
if hap > 128 :
	print("\n이전회차는 고저값이 높습니다.")
	
if hap < 128 :
	print("\n이전회차는 고저값이 낮습니다.")
	
	

dnots=sorted(dnot_dict.items(), key=operator.itemgetter(1), reverse=True )
print("\n미출현 번호 목록 (내림차순)\n",dnots)

dnots_list=[]
for i in range(45):
	ia=dnots[i]
	ia=ia[0]
	dnots_list.append(ia)
dnots_favorit=dnots_list[0:22]


dnots_favorit=[]
for item in range(23):
	ia=dnots[item]
	ia=ia[0]
	dnots_favorit.append(ia)

# 양력 , 음력 변수를 도입

if int(date) > 0:
	print("\n날짜 적용이 되었습니다.")
	print("양력 :", date[0],date[1],"월",date[2],date[3],"일")
	print("음력 :", date[4],date[5],"월",date[6],date[7],"일")

	for i in range(135):
		choice_from_date=random.randint(0,7)
		if choice_from_date > 0:
			choice_1num=date[choice_from_date]
			choice=0
			
			while choice == 0:
				choice_lot_num=random.randint(1,45)
				choice_lot_num=choice_lot_num
				cln_str=str(choice_lot_num)
				for cln_str in str(choice_1num):
					choice=choice_lot_num
			add(choice)
			

print("\n사전 조합 개수 자동 설정... 총 ",avera,"개로 계산됨.")


		

# 기초 확률 순위 계산


ranlists_dict={}
for i in range(45):
	ai=i+1
	aic=ranlists.count(ai)
	ranlists_dict[ai]=aic
sort_of_ranlists_dict=sorted(ranlists_dict.items(), key=operator.itemgetter(1), reverse=True )


# 본격 사전 조합 생성 함수 

for i in range(3):
	shuffle(ranlists)

def make_num():
	r=random.choice(ranlists)
	if r in mlist:
		make_num()		
	for i in range(1):
		ranlists.append(r)	
	return r

def make_mlist():
	global mlist
	mlist=[]
	for i in range(6):
		r=make_num()
		mlist.append(r)
	miner_nums=list(set(olist)-set(mlist))
	for i in range(len(miner_nums)):
		sel_num=miner_nums[i]	
		for i in range(2):
			ranlists.append(sel_num)
	mlist=list(set(mlist))
	if len(mlist) < 6:
		make_mlist()
	return mlist
	

rcs=[]
rcss=[]

proc_avera=int(avera/cc)
numproc=range(0,cc)
print("\n기초 순위\n",sort_of_ranlists_dict)
if cc > 1:
	print("\n귀하의 PC의 코어가",cc,"개 로 확인 되었습니다. 멀티프로세싱을 활용 하여 연산 합니다...\n")
	ac=int(cc/2)
else:
	ac=1

# 멀티 프로세싱 기반의 연산 작업 시작

os.system('cat /dev/null > /tmp/randlot_averas')
def proc_make(numproc):
	file=open('/tmp/randlot_averas' , 'a')
	for i in range(proc_avera):
		mlist=make_mlist()
		mlist.sort()
		data=str(mlist)
		data=data.strip('[]')
		file.write(data+"\n")
	file.close()

if __name__ == '__main__':
	#멀티 쓰레딩 Pool 사용
	pool = multiprocessing.Pool(processes=cc) # 현재 시스템에서 사용 할 프로세스 개수
	pool.map(proc_make, numproc)
	pool.close()
	pool.join()


print("다중연산이 끝나고 정리를 시작합니다...")

with open('/tmp/randlot_averas', 'r') as f:
	list_file = f.readlines()
alist = [line.rstrip('\n') for line in list_file]

rcs=[]
for i in range(len(alist)):
	ai=alist[i]
	dot=','
	ai=ai.split(dot)
	ai=list(map(int, ai))
	chk_ai=list(set(ai))
	if len(chk_ai) == 6:
		rcs.append(ai)




for i in range(cc):
	
	shuffle(rcs)


# 연산된 난수 총 순위

ranlists=[]
for i in range(len(rcs)):
	ai=rcs[i]
	for i in range(6):
		ri=ai[i]
		ranlists.append(ri)
		
ranlists_dict={}
for i in range(45):
	ai=i+1
	aic=ranlists.count(ai)
	ranlists_dict[ai]=aic
sort_of_coumputed_list=sorted(ranlists_dict.items(), key=operator.itemgetter(1), reverse=True )


favorit_list=[]
for i in range(23):
	favorit=sort_of_ranlists_dict[i]
	favorit=favorit[0]
	favorit_list.append(favorit)

low_list=[]
for i in range(22):
	seven=sort_of_ranlists_dict[-i]
	seven=seven[0]
	low_list.append(seven)

seven_list=[]
for i in range(7):
	seven=sort_of_ranlists_dict[i]
	seven=seven[0]
	seven_list.append(seven)


ranlists_average=len(ranlists)/45
ranlists_dict_average={}
for i in range(45):
	ai=i+1
	ranlists_dict_average[ai]=(ranlists_dict[ai]-ranlists_average)/3.14/45+ranlists_average
	

after_flist=[]
for i in range(23):
	favorit=sort_of_coumputed_list[i]
	favorit=favorit[0]
	after_flist.append(favorit)

ranlists=[]
for i in range(45):
	ai=i+1
	wr=ranlists_dict_average[ai]
	wr=int(wr)
	for i in range(wr):
		ranlists.append(ai)

	
computed_favorit_list=[]
for i in range(23):
	seven=sort_of_coumputed_list[i]
	seven=seven[0]
	computed_favorit_list.append(seven)


rate1=0

if test == "test":
	print("\n테스트가 옵션에 적용 되어 있어 테스트 연산을 시작 합니다...")

	after_num=nnum+1
	load_wb = load_workbook("/tmp/randlot_after_data.xlsx", data_only=True)
	sheet=load_wb.active
	sheet.title = "1"
	load_ws = load_wb["1"]

	for i in range(1):
		a=4+i
		datasel1="B"+str(a)
		data=load_ws[datasel1].value
		data1=int(data)

		datasel2="N"+str(a)
		data=load_ws[datasel2].value
		data2=int(data)

		datasel3="O"+str(a)
		data=load_ws[datasel3].value
		data3=int(data)


		datasel4="P"+str(a)
		data=load_ws[datasel4].value
		data4=int(data)

		datasel5="Q"+str(a)
		data=load_ws[datasel5].value
		data5=int(data)


		datasel6="R"+str(a)
		data=load_ws[datasel6].value
		data6=int(data)

		datasel7="S"+str(a)
		data=load_ws[datasel7].value
		data7=int(data)
		after_list=[data2]+[data3]+[data4]+[data5]+[data6]+[data7]

		if a ==4:
			print("\n테스트 적용 회차: ", after_num, " 의 당첨번호는 ",after_list,"입니다.")

	after_list=set(after_list)

# 각 연산 전 연산 후 상위 리스트 

top_of_sort_of_ranlists_dict=[]
for i in range(23):
	li=sort_of_ranlists_dict[i]
	top_of_sort_of_ranlists_dict.append(li)

top_sort_of_coumputed_list=[]
for i in range(7):
	li=sort_of_coumputed_list[i]
	top_sort_of_coumputed_list.append(li)


hists=[]

rate_per=0
hist=[]
dnot_hist_list=olist
firsts=[]
lasts=[]
three=[]
brc=nrc=random.choice(rcs)
rcns=[]
not_list=list(range(1,46))


# 중복 번호 찾아내기  (match point)

print("\nmatch 포인트를 연산 합니다...")



os.system('cat /tmp/randlot_averas |sort |uniq -c -d |grep "4 " > /tmp/randlot_matchs')
os.system('cat /tmp/randlot_averas |sort |uniq -c -d |grep "3 " >> /tmp/randlot_matchs')
os.system('cat /tmp/randlot_averas |sort |uniq -c -d |grep "2 " >> /tmp/randlot_matchs')

os.system('sed -i "s/3 //g" /tmp/randlot_matchs')
os.system('sed -i "s/4 //g" /tmp/randlot_matchs')
os.system('sed -i "s/2 //g" /tmp/randlot_matchs')



with open('/tmp/randlot_matchs', 'r') as f:
	list_file = f.readlines()
alist = list([line.rstrip('\n') for line in list_file]) 


matchs=[]
for i in range(len(alist)):
	ai=alist[i]
	dot=','
	ai=ai.split(dot)
	ai=list(map(int, ai))
	chk_ai=list(set(ai))
	if len(chk_ai) == 6:
		for i in range(6):
			ia=ai[i]
			matchs.append(ia)



print("\n매치 포인트 연산 내림차순")

counter_matchs=collections.Counter(matchs)
print(counter_matchs.most_common(),"\n")

matchs_list=counter_matchs.most_common(23)

matchs_favorit_list=[]
for i in range(len(matchs_list)):
	ia=matchs_list[i]
	num=ia[0]
	matchs_favorit_list.append(num)
 

print("\n",matchs_favorit_list)

print("\n추천 번호 출력")


# 번호 뽑기
oldlist=[]
rcsd=[]
for i in range(amount):
	if len(not_list) < 1 :
		not_list=list(range(1,46))
		
	selnice=random.randint(0,5)
	chk_dnots_rcs=0
	sel_rc=random.choice(rcs)


	while True:
		if 0 < len(set(sel_rc) & set(favorit_list)):
			if 0 < len(set(sel_rc) & set(dnots_favorit)) :
				if 0 < len(set(sel_rc) & set(computed_favorit_list)):
						if 0 < len(set(sel_rc) & set(onums)):
							if 0 < len(set(sel_rc) & set(not_list)):
								if 0 < len(set(sel_rc) & set(matchs_favorit_list)):

									if len(oldlist) > 11:
										if 0 < len(set(sel_rc) & set(oldlist)):
											rc=sel_rc
											break
									else:
										rc=sel_rc
										break
							
		sel_rc=random.choice(rcs)
		
	for i in range(6):
		rcsd.append(rc[i])
		oldlist.append(rc[i])

	rc.sort()
	oldlist=list(set(oldlist))
	if len(oldlist) > 23:
		oldlist=[]

	miner_nums=list(set(olist)-set(rc))
	for i in range(len(miner_nums)):
		sel_num=miner_nums[i]	
		
	brc=rc
	rcns.append(rc)
	for i in range(6):
		if rc[i] in not_list:
			not_list.remove(rc[i])

	if rc in not_list:
		not_list.remove(i)
	
	def sel_for_rc():
		sel=random.choice(rc)
		if sel == 0:

			sel_for_rc()
		return sel

	if half == 1:
		for i in range(2):

			rc.remove(sel_for_rc())
			rc.append(0)
			
	if test =="test":
		test_rc=set(after_list) & set(rc)
		if len(test_rc) > 2:
			rated=len(test_rc)*0.25
			rate_per=rate_per+rated
		print(rc, test_rc, len(test_rc), )
	if test != "test":
		print(rc)

# 테스트 출력

os.system('cat /dev/null > /tmp/randlot_chklist')


if test =="test":
	check_top_nums=set(after_list) & set(favorit_list)

	per_fav=len(check_top_nums)/23 *100
	print("\n기초 확률 상위 적중 확률 :",per_fav,"%")
	print(check_top_nums)

	check_top_nums=set(after_list) & set(computed_favorit_list)
	per_fav=len(check_top_nums)/23 *100
	print("\n연산 상위 적중 확률 :",per_fav,"%")
	print(check_top_nums)

	check_top_nums=set(after_list) & set(matchs_favorit_list)
	per_fav=len(check_top_nums)/23 *100
	print("\n매치 포인트 적중 확률 :",per_fav,"%")
	print(check_top_nums)


	a3=0
	a4=0
	a5=0
	a6=0
	for i in range(len(rcs)):
		rcs_list = rcs[i]
		chk_rcs_list=set(rcs_list) & set(after_list)
		
		if len(chk_rcs_list) == 3:
			a3=a3+1
		if len(chk_rcs_list) == 4:
			a4=a4+1
		if len(chk_rcs_list) == 5:
			a5=a5+1	
			chkdata=str(rcs_list)
			f = open('/tmp/randlot_chklist', 'a')
			f.write(str(chkdata)+"\n")
		if len(chk_rcs_list) == 6:
			a6=a6+1
	print("전체 연산된 리스트", len(rcs),"가지 경우 중... 3개 적중 :",a3," 4개 적중 :",a4,"  5개 적중 :",a5," 6개 적중 :",a6)

if test =="test":
	rate_per=rate_per/amount*100
	print("\n총 출력 된 추천 번호들의 총 당첨 확률 :",rate_per,"%")
	a3=0
	a4=0
	a5=0
	a6=0
	for i in range(len(rcns)):
		rcs_list=rcns[i]
		chk_rcs_list=set(rcs_list) & set(after_list)
		
		if len(chk_rcs_list) == 3:
			a3=a3+1
		if len(chk_rcs_list) == 4:
			a4=a4+1
		if len(chk_rcs_list) == 5:
			a5=a5+1	
		if len(chk_rcs_list) == 6:
			a6=a6+1
	print("전체 출력된 추천 리스트", len(rcns),"개 중... 3개 적중 :",a3," 4개 적중 :",a4,"  5개 적중 :",a5," 6개 적중 :",a6)




etime=time.time()
proc_time=etime-stime

not_lists= set(olist) - set(rcsd)


print("비출현 번호 : ", list(not_lists))

print("\n연산에 걸린 총 시간은", int(proc_time) , "초 입니다.")

